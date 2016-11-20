#coding：utf-8

from openpyxl import load_workbook

def getaddresslist(filename):
    addresslist=[]
    wb=load_workbook(filename)
    ws=wb.active
    for column in range(1,ws.max_column+1):
        if ws.cell(row=1,column=column).value=='住址':
            for row in range(2,ws.max_row+1):
                addresslist.append([ws.cell(row=row,column=column).value])
            break
    return addresslist

