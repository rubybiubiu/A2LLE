#coding：utf-8

from openpyxl import load_workbook
#from A2L import BaiDuMapAPI
def DealWithExcel(filename):
    from A2L import BaiDuMapAPI
    wb=load_workbook(filename)
    ws=wb.active
    ws.title='address'
    ws['C1']='经度'
    ws['D1']='纬度'

    for row in range(2,ws.max_row+1):
        address=ws['B'+str(row)].value
        x,y,z=BaiDuMapAPI.address2LngLat(address)
        if z==0:
            ws['C' + str(row)].value=x
            ws['D' + str(row)].value=y
        else:
            ws['C' + str(row)].value=z
            ws['D' + str(row)].value=z

    wb.save(filename)





