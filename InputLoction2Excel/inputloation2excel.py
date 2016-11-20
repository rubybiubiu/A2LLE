#coding：utf-8

from openpyxl import load_workbook

def inputloction2excel(filename,locationlist):
    wb = load_workbook(filename)
    ws = wb.active  # 激活第一个表格
    ws.cell(row=1,column=ws.max_column+1).value='经度'
    ws.cell(row=1,column=ws.max_column+1).value='纬度'
    i=2
    for x in locationlist:
        ws.cell(row=i,column=ws.max_column-1).value=x[1]
        ws.cell(row=i, column=ws.max_column).value = x[2]
        i+=1
    wb.save(filename)


